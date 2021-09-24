from re import M
import openpyxl


book = openpyxl.load_workbook(r"C:\Users\miwan\Desktop\Everest Times.xlsx")

sheet = book["Sheet1"]

print(sheet["C6"].value)
print(sheet.cell(3,4).value)

sheet["B23"] = "Hello There"

book.save(r"C:\Users\miwan\Desktop\Everest Times.xlsx")

for row in sheet.iter_rows(values_only=True):
    print(row)